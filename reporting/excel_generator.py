"""
Generates a formatted Excel (.xlsx) file from processed leads.
"""
import logging
from datetime import date
from pathlib import Path
from typing import List, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

EXPORTS_DIR = Path("data/exports")

# Column definitions: (header, lead_key, width)
COLUMNS = [
    ("Business Name",  "name",        35),
    ("Phone Number",   "phone",       18),
    ("Location",       "address",     35),
    ("Category",       "niche",       18),
    ("Score",          "score",        8),
    ("Lead Type",      "_lead_type",  12),
    ("Opportunity",    "opportunity", 45),
    ("Hook",           "hook",        50),
]

HEADER_FILL   = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
HOT_FILL      = PatternFill("solid", fgColor="FFE5E5")
WARM_FILL     = PatternFill("solid", fgColor="FFF9E5")
BORDER_SIDE   = Side(style="thin", color="CCCCCC")
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                        top=BORDER_SIDE, bottom=BORDER_SIDE)


def _lead_type(lead: Dict) -> str:
    intent = lead.get("intent", "")
    score  = int(lead.get("score") or 0)
    if intent == "HIGH" or score >= 8:
        return "HOT"
    if intent == "MEDIUM" or score >= 6:
        return "WARM"
    return "TEST"


def generate_excel(leads: List[Dict], filename: str = None) -> Path:
    """
    Create a formatted Excel file from leads list.
    Returns the Path to the saved file.
    """
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    today = date.today().strftime("%Y-%m-%d")
    path  = EXPORTS_DIR / (filename or f"nexora_leads_{today}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # ── Header row ─────────────────────────────────────────────────────────────
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = CELL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── Data rows ──────────────────────────────────────────────────────────────
    for row_idx, lead in enumerate(leads, start=2):
        ltype = _lead_type(lead)
        fill  = HOT_FILL if ltype == "HOT" else (WARM_FILL if ltype == "WARM" else None)

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            value = lead.get("_lead_type", ltype) if key == "_lead_type" else lead.get(key, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = CELL_BORDER
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 40

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    hot  = sum(1 for l in leads if _lead_type(l) == "HOT")
    warm = sum(1 for l in leads if _lead_type(l) == "WARM")

    summary_data = [
        ("Date",         today),
        ("Total Leads",  len(leads)),
        ("HOT Leads",    hot),
        ("WARM Leads",   warm),
        ("City",         leads[0].get("city", "") if leads else ""),
    ]
    for r, (k, v) in enumerate(summary_data, start=1):
        ws2.cell(row=r, column=1, value=k).font  = Font(bold=True)
        ws2.cell(row=r, column=2, value=str(v))
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 20

    wb.save(path)
    logger.info(f"Excel file saved: {path} ({len(leads)} leads)")
    return path
